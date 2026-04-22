"""
build_spec.py
Extracts ALL knowledge from Sharp VE BLDC 11/13kg Excel spec file
into a single structured sharp_spec.json knowledge base.
Run once: python build_spec.py
"""
import json
import openpyxl
import datetime

FILE = 'Sharp VE BLDC 11,13kg V0.xlsx'
wb   = openpyxl.load_workbook(FILE, data_only=True)

def val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, datetime.timedelta):
        return int(v.total_seconds())
    return v

def rows(ws, max_row=60):
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        clean = []
        for c in row:
            if isinstance(c, datetime.timedelta):
                clean.append(int(c.total_seconds()))
            else:
                clean.append(c)
        if any(c is not None for c in clean):
            out.append({"row": i, "values": [str(c) if c is not None else None for c in clean[:25]]})
    return out

spec = {}

# ── 1. META ──────────────────────────────────────────────────────────────────
spec["meta"] = {
    "source_file": FILE,
    "model": "Sharp VE BLDC 11/13 kg",
    "built_at": str(datetime.datetime.now()),
    "total_programs": 12
}

# ── 2. PROGRAM LIST ───────────────────────────────────────────────────────────
spec["programs"] = {
    "list": [
        "Regular", "Quick", "Heavy", "Baby Care",
        "Cotton", "Delicates", "Wool", "Jeans",
        "Blanket", "Quick Rinse", "Sports Wear", "Tub Clean"
    ],
    "vs_ac_extra": ["Quick Rinse", "Sports Wear"],
    "note": "BLDC has 12 programs vs AC 10 programs (Differences Summary sheet)"
}

# ── 3. COURSE GROUPS ──────────────────────────────────────────────────────────
spec["course_groups"] = {
    "Group 1": {
        "programs": ["Regular", "Quick", "Baby Care", "Quick Rinse"],
        "note_quick": "Quick skips M2 and M3 motion phases",
        "source": "Course Group 1 sheet row 2-3"
    },
    "Group 2": {
        "programs": ["Jeans", "Cotton", "Heavy"],
        "source": "Course Group 2 sheet row 2"
    },
    "Group 3": {
        "programs": ["Wool", "Delicates", "Sports Wear"],
        "source": "Course Group 3 sheet row 2"
    },
    "Special": {
        "programs": ["Blanket", "Tub Clean", "Fragrance Rinse Spin"]
    }
}

# ── 4. SEQUENCE CHART ─────────────────────────────────────────────────────────
ws_seq = wb['Sequence Chart']

# Read program columns from row 3
prog_row = [ws_seq.cell(3, c).value for c in range(1, 46)]
level_row = [ws_seq.cell(4, c).value for c in range(1, 46)]

# Build column index: program -> list of (level, col_index)
prog_cols = {}
current_prog = None
for ci, (p, lv) in enumerate(zip(prog_row, level_row), 1):
    if p and p not in ('Process', 'Selectable'):
        current_prog = p
        prog_cols.setdefault(current_prog, [])
    if current_prog and lv and 'LEV' in str(lv):
        prog_cols[current_prog].append((str(lv), ci))

# Helper: read a row value for a program/level
def seq_val(program, level, excel_row):
    entries = prog_cols.get(program, [])
    for lv, ci in entries:
        if lv == level:
            v = ws_seq.cell(excel_row, ci).value
            if isinstance(v, datetime.timedelta):
                return int(v.total_seconds())
            return v
    return None

programs_in_seq = list(prog_cols.keys())
sequence_data = {}
for prog in programs_in_seq:
    sequence_data[prog] = {}
    for lv, _ in prog_cols[prog]:
        sequence_data[prog][lv] = {
            "main_wash_sec":    seq_val(prog, lv, 14),
            "rinse_count":      seq_val(prog, lv, 19),
            "drain_sec":        seq_val(prog, lv, 20),
            "water_fill_sec":   seq_val(prog, lv, 30),
            "rinse_wash_sec":   seq_val(prog, lv, 34),
            "balance_spin_sec": seq_val(prog, lv, 24),
            "final_spin_sec":   seq_val(prog, lv, 44),
            "inertia_stop_sec": seq_val(prog, lv, 48),
            "anti_wrinkle_sec": seq_val(prog, lv, 50),
        }

spec["sequence_chart"] = sequence_data

# ── 5. WATER LEVELS ────────────────────────────────────────────────────────────
spec["water_levels"] = {
    "LEV-1": {"display": 1, "liters": 28, "load_kg": "0-2"},
    "LEV-2": {"display": 2, "liters": 50, "load_kg": "3-5"},
    "LEV-3": {"display": 3, "liters": 71, "load_kg": "6-8"},
    "LEV-4": {"display": 4, "liters": 93, "load_kg": "9-13"},
    "formula": "y = 21.6x + 6.5 (x=load kg, y=water liters)",
    "source": "Water Levels sheet"
}

# ── 6. MOTOR TIMINGS PER LEVEL (Course Groups 1/2/3) ─────────────────────────
motor_timings = {}
for grp_name, sheet_name in [("Group1","Course Group 1"),
                               ("Group2","Course Group 2"),
                               ("Group3","Course Group 3")]:
    try:
        ws_g = wb[sheet_name]
        motor_timings[grp_name] = {}
        for r in range(8, 13):
            lv = ws_g.cell(r, 2).value
            if lv and str(lv).isdigit():
                cw  = ws_g.cell(r, 12).value
                ccw = ws_g.cell(r, 14).value
                motor_timings[grp_name][str(lv)] = {
                    "m2_cw_sec":  float(cw)  if cw  else 0.5,
                    "m2_ccw_sec": float(ccw) if ccw else 0.5
                }
    except Exception as e:
        motor_timings[grp_name] = {"error": str(e)}

spec["motor_timings"] = motor_timings

# ── 7. ERRORS ─────────────────────────────────────────────────────────────────
ws_err = wb['Errors']
errors = []
for r in range(3, 25):
    name = ws_err.cell(r, 2).value
    code = ws_err.cell(r, 3).value
    detection = ws_err.cell(r, 4).value
    release   = ws_err.cell(r, 5).value
    remedy    = ws_err.cell(r, 6).value
    if code and str(code).strip() not in ('', 'nan', 'Error Code'):
        errors.append({
            "code": str(code).strip(),
            "name": str(name).strip() if name else "",
            "detection": str(detection)[:200] if detection else "",
            "release": str(release)[:200] if release else "",
            "daq_detectable": str(code).strip() in ["E1","E2","E5","E7-4","Eb-1","E7-1","E7-2","E7-3","E6-1","EA"]
        })
spec["errors"] = errors

# ── 8. SPIN CONTROL SPECS ──────────────────────────────────────────────────────
spec["spin_control"] = {
    "standard_programs": {
        "note": "Regular, Heavy, Cotton, Baby Care, Jeans, Quick Rinse, Blanket",
        "rpm_curve": [
            {"time_sec": 0,   "rpm": 0},
            {"time_sec": 15,  "rpm": 300},
            {"time_sec": 35,  "rpm": 300},
            {"time_sec": 155, "rpm": 600},
            {"time_sec": 160, "rpm": 600},
            {"time_sec": 180, "rpm": 700},
            {"time_sec": 240, "rpm": 700}
        ],
        "max_rpm": 700
    },
    "gentle_programs": {
        "note": "Delicates, Wool, Sports Wear",
        "rpm_curve": [
            {"time_sec": 0,   "rpm": 0},
            {"time_sec": 15,  "rpm": 300},
            {"time_sec": 35,  "rpm": 300},
            {"time_sec": 155, "rpm": 400},
            {"time_sec": 240, "rpm": 400}
        ],
        "max_rpm": 400
    },
    "source": "Spin Control Specs sheet"
}

# ── 9. PUMP CONTROL SPECS ─────────────────────────────────────────────────────
spec["pump_control"] = {
    "max_continuous_sec": 150,
    "required_off_sec": 10,
    "overflow_threshold_liters": 103,
    "operation_mode": "Continuous ON",
    "e1_timeout_min": 15,
    "note": "After 150s continuous ON, pump must stop for 10s. E1 triggers after 15 min no drain.",
    "source": "Pump Control Specs sheet"
}

# ── 10. WEIGHT DETECTION ──────────────────────────────────────────────────────
spec["weight_detection"] = {
    "mechanism": "RPM sensor",
    "start_direction": "CCW first",
    "sequence": "CCW_ON → CCW_OFF → CW_ON → CW_OFF (repeated 4 times)",
    "cw_on_ms": 300,
    "cw_off_ms": 600,
    "ccw_on_ms": 300,
    "ccw_off_ms": 600,
    "total_repeats": 4,
    "cancelled_when": "Water level manually set OR water > LEV-1 exists",
    "source": "Weight Detection Specs sheet"
}

# ── 11. DOOR OPENING SPECS ────────────────────────────────────────────────────
spec["door_specs"] = {
    "E2_trigger_delay_sec": 0.2,
    "allowed_open_during": "Water filling before wash or rinse ONLY",
    "after_close": "Machine continues automatically without pressing pause",
    "source": "Door Opening Specs sheet"
}

# ── 12. CHILD LOCK SPECS ──────────────────────────────────────────────────────
spec["child_lock"] = {
    "activation": "Press Soak + Delay Start simultaneously during execution",
    "case_A": {
        "condition": "Door opened then closed before 20 seconds",
        "result": "Washing continues normally - PASS"
    },
    "case_B": {
        "condition": "Door still open after 20 seconds",
        "result": "Pump activates immediately, drains to reset level, WM turns off after 10 min"
    },
    "buzzer": "Every 10 minutes, buzzer alarms for 10 seconds during error",
    "source": "Child Lock Specs sheet"
}

# ── 13. WATER INLET ───────────────────────────────────────────────────────────
spec["water_inlet"] = {
    "cold": "Cold valve only",
    "warm": "Cold + Hot valves (cold fills to 50%, then hot fills to target)",
    "hot": "Cold + Hot valves (cold fills to 33%, then hot fills to target)",
    "rinse_temp": "Cold only always",
    "error_EE4_threshold_deg": 63,
    "source": "Water Inlet sheet"
}

# ── 14. MEMS SPECS ────────────────────────────────────────────────────────────
spec["mems"] = {
    "function": "Unbalance detection during spin",
    "axes": {"X": "Up/Down", "Y": "Left/Right", "Z": "Front/Rear"},
    "border_values_regular": {"X": 10, "Y": 28, "Z": 28, "NH": 40, "CH": 5},
    "border_values_factory": {"X": 10, "Y": 38, "Z": 38, "NH": 60, "CH": 20},
    "source": "Mems Specs sheet"
}

# ── 15. POWER CUT OFF ─────────────────────────────────────────────────────────
spec["power_cut"] = {
    "behavior": "WM saves data and continues from same point when power returns",
    "exception": "If cut during wash/rinse, restarts from beginning of that motion",
    "resume": "Operates immediately when power returns - no button press needed",
    "source": "Power Cut off sheet"
}

# ── 16. SOAK ─────────────────────────────────────────────────────────────────
spec["soak"] = {
    "available_for": ["Regular", "Wool", "Blanket"],
    "not_allowed_when": "Wash time = 3 min",
    "durations": ["1 hour", "2 hours", "4 hours"],
    "soak_motor": {"CW_sec": 2.4, "OFF_sec": 2.6, "CCW_sec": 2.4, "OFF2_sec": 2.6},
    "source": "Soak sheet"
}

# ── 17. BLANKET ──────────────────────────────────────────────────────────────
spec["blanket"] = {
    "capacity_kg": 3,
    "water_level": "LEV-4",
    "wash_cw_sec": 3.8,
    "wash_ccw_sec": 3.8,
    "total_wash_sec": 710,
    "source": "Blanket sheet"
}

# ── 18. OPTIONS (Course Option sheet) ─────────────────────────────────────────
spec["course_options"] = {
    "selectable": ["Water Temp", "Wash Time", "Rinse Times", "Spin Time", "Water Level", "Soak", "Delay Start", "Anti-Wrinkle"],
    "manual_wash_times_min": [3, 9, 12, 18],
    "manual_rinse_times": [1, 2, 3],
    "manual_spin_times": ["1 min", "5 min", "9 min", "Super Spin 20 min"],
    "water_levels": [1, 2, 3, 4],
    "source": "Course Option sheet"
}

# ── 19. LAYOUT (UI panel) ─────────────────────────────────────────────────────
spec["ui_layout"] = {
    "programs_count": 12,
    "options": ["Anti-Wrinkle", "Delay Start", "Soak", "Child Lock"],
    "manual_selections": [
        "3 Water Temps: COLD/WARM/HOT",
        "4 Wash Times: 3/9/12/18 min",
        "3 Rinse Times: 1/2/3 times",
        "4 Spin Times: 1/5/9 min + Super Spin",
        "4 Water Levels: 1/2/3/4"
    ],
    "source": "Layout sheet"
}

# ── 20. DIFFERENCES SUMMARY ────────────────────────────────────────────────────
spec["differences_summary"] = {
    "BLDC_vs_AC": {
        "programs": "BLDC=12, AC=10",
        "extra_in_BLDC": ["Quick Rinse", "Sports Wear"],
        "display": "BLDC=3 digits, AC=4 digits",
        "balance_spin": "BLDC=Spin Curve, AC=ON/OFF"
    },
    "source": "Differences Summary sheet"
}

# ── SAVE ─────────────────────────────────────────────────────────────────────
with open('sharp_spec.json', 'w', encoding='utf-8') as f:
    json.dump(spec, f, indent=2, ensure_ascii=False, default=str)

print(f"sharp_spec.json created successfully!")
print(f"Sections: {list(spec.keys())}")
print(f"Programs: {spec['programs']['list']}")
print(f"Errors captured: {len(spec['errors'])}")
print(f"Sequence programs: {list(spec['sequence_chart'].keys())}")
